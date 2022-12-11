# -*- coding: utf-8 -*-
"""
Determining amino acid substitution odds-ratio proportions
based on expected and observed values
"""
from Sampled_confidence_intervals import confidence_interval
from Sampled_confidence_intervals import multidimensional_shifting
import openpyxl as oxl
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
# Variables that have to be changed
fileName = "begin_to_omicron.xlsx"
masterSheet = "masterSheet"
total_sequences = 7599578
samples = 100 # Number of samples taken, >1000 recommended
samp_size = 5 * total_sequences
mutationList = ['R346T',
                'R346K',
                'K444T',
                'K444R',
                'K444N',
                'L452R',
                'L452M',
                'L452Q',
                'N460K',
                'F486S',
                'F486V',
                'F490V',
                'F490L',
                'R493Q',
                'S494P']
# Values dictionary that will store the mutations as keys and the confidence interval/median as the values
excel_values = {}
# Excel sheet variable
xlsx = oxl.load_workbook(fileName)

# Find observed values for co-occurrence of mutations
# mutation1 will be the sheet name and mutation2 will be the compared mutation
# Returns the number of observed co-occurrences
def find_cooccurrence_count(mutation1, mutation2):
    for row in xlsx[mutation1].iter_rows(min_row=1, min_col=1, values_only = True):
        # Iterates through column2 in the row
        modMutation = "S:" + mutation2
        if row[1] == modMutation:
            return row[6] # Returns number co-occurrences of mutation 1 and mutation 2


def find_mutation_count(mutation):
    for row in xlsx[masterSheet].iter_rows(min_row=1, min_col=1, values_only = True):
        modMutation = "S:" + mutation
        if row[0] == modMutation:
            return row[5]
        
# Nested for loop interates through the mutation list twice and fills mutation dictionary values
for mutation1 in mutationList:
    for mutation2 in mutationList:
        if mutation1[1:4] == mutation2[1:4]: # Prevents mutation calculations if the position is the same
            break
        else:
            sortedMutationPair = ''.join(sorted([mutation1, mutation2])) # Creates a sorted string to prevent redundancy
            mutationCombination = mutation1 + " and " + mutation2 # Creates a key for the dictionary with the mutations names as the key
            # Attempts to find the co-occurrence count
            try:
                mutationAttributes = [find_mutation_count(mutation1), find_mutation_count(mutation2), find_cooccurrence_count(mutation1, mutation2)] 
                excel_values[mutationCombination] = mutationAttributes # Adds mutation attributes to the dictionary
            # If there is no data, return Data_not_found to the dictionary
            except:
                excel_values[mutationCombination] = 'Data_not_found'
                
mutationAttributeDict = {} # Creates dictionary to hold mutation attributes
numpyConversion = np.array(list(excel_values.items()))       

""" Dictionary-based function
for key in excel_values:
    mutationStats = [] # Temp list to hold mutation values
    
    # PASSING DATA TO FUNCTION

    if isinstance(excel_values[key][0], int) and isinstance(excel_values[key][1], int) and isinstance(excel_values[key][2], int):
        # print(excel_values[key][0], excel_values[key][1], excel_values[key][2])
        low, med, high = confidence_interval(excel_values[key][0], excel_values[key][1], excel_values[key][2], 
                                             total_sequences, samples, samp_size, confidence=0.95, draw_result=False)
        mutationStats = [low, med, high]
        mutationAttributeDict[key] = mutationStats
"""

# Numpy based function
for i in range(len(numpyConversion)):
    mutationStats = [] # Temp list to hold mutation values
    
    # Checks that values in array are integers
    if isinstance(numpyConversion[i][1][0], int) and isinstance(numpyConversion[i][1][1], int) and isinstance(numpyConversion[i][1][2], int):
        low, med, high = confidence_interval(numpyConversion[i][1][0], numpyConversion[i][1][1], numpyConversion[i][1][2], 
                                             total_sequences, samples, samp_size, confidence=0.95, draw_result=False)
        mutationStats = [low, med, high] # Updates temporary list to hold confidence interval values
        mutationAttributeDict[numpyConversion[i][0]] = mutationStats
        

# Create error plot
# Creates variable lists for plot
keyList = [i for i in mutationAttributeDict]
valueMedian = [mutationAttributeDict[i][1] for i in mutationAttributeDict]
valueYErr = [(mutationAttributeDict[i][2]-mutationAttributeDict[i][0])/2 
             for i in mutationAttributeDict]

# Sorts list in order
sortableList = [[valueMedian[i], keyList[i], valueYErr[i]] for i in range(len(keyList))]
sortedList = sorted(sortableList)
sortedMedian = []
sortedKeyList = []
sortedValueYErr = []

# Creates sorted lists of each attribute, if need reversed, replace with 
# for i in range(0, len(sortedList), -1)
for i in range(len(sortedList)):
    sortedMedian.append(sortedList[i][0])
    sortedKeyList.append(sortedList[i][1])
    sortedValueYErr.append(sortedList[i][2])

#Creates the plot
fig,ax = plt.subplots(figsize=(20,12)) # Creates size of Plot
sns.set(style= "whitegrid")
ax.errorbar(sortedKeyList, sortedMedian, yerr = sortedValueYErr, capsize = 10, fmt = 'o')
plt.xticks(rotation=60) # Changes the x-axis label orientation
fig.show()


""" try-except alternative if the if isinstance portion does not produce adequate values
    print(key)
    mutationStats = [] # Temp list to hold mutation values
    try: #Attempts to use Sampled_confidence_intervals function
        print(excel_values[key][0])
        print(excel_values[key][1])
        print(excel_values[key][2])
        low, med, high = confidence_interval(excel_values[key][0], excel_values[key][1], excel_values[key][2], 
                                             total_sequences, samples, samp_size, confidence=0.95, draw_result=False)
        mutationStats = [low, med, high]
        mutationAttributeDict[key] = mutationStats
    except: #If function cannot be run, skip this iteration
        continue
"""


